"""Polars-first Chronicle Android raw data preprocessing orchestration."""

from __future__ import annotations

import logging
import multiprocessing
import time
from collections.abc import Callable
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Any

import polars as pl

from chronicle_preprocessing_app.config.constants import (
    GAP_TIMESTAMPS_SIDECAR_SUFFIX,
    PLOTTED_FOLDER_SUFFIX,
    PREPROCESSED_FILE_SUFFIX,
    PREPROCESSED_FOLDER_SUFFIX,
    Column,
    InteractionType,
)
from chronicle_preprocessing_app.core.config import PreprocessingOptions, ProcessingStats
from chronicle_preprocessing_app.core.preprocessing.app_filter_preprocessor import (
    AppFilterPreprocessor,
)
from chronicle_preprocessing_app.core.preprocessing.app_usage_preprocessor import (
    AppUsagePreprocessor,
    NoAppUsageDataError,
)
from chronicle_preprocessing_app.core.preprocessing.column_preprocessor import (
    ColumnPreprocessor,
)
from chronicle_preprocessing_app.core.preprocessing.polars_fast_path import (
    PolarsFastPathPreprocessor,
    supports_polars_fast_path,
)
from chronicle_preprocessing_app.core.preprocessing.screen_usage_preprocessor import (
    ScreenUsagePreprocessor,
)
from chronicle_preprocessing_app.core.preprocessing.study_date_provider import (
    StudyDateRangeProvider,
)
from chronicle_preprocessing_app.core.preprocessing.timestamp_preprocessor import (
    TimestampPreprocessor,
)
from chronicle_preprocessing_app.core.preprocessing.timezone_preprocessor import (
    TimezonePreprocessor,
)
from chronicle_preprocessing_app.utils.file_utils import (
    get_matching_files_from_folder,
    read_app_codebook,
    read_apps_forcing_screen_open_file,
    read_background_apps_file,
    read_filter_file,
)

LOGGER = logging.getLogger(__name__)

# Columns that must be present in any raw Chronicle Android CSV before
# preprocessing can proceed.  These are the columns the pipeline reads from the
# raw file — not derived/output columns like start_timestamp or
# possible_device_model that the pipeline generates itself. timezone is
# optional: missing row timezones use the documented UTC fallback.
REQUIRED_RAW_COLUMNS: frozenset[str] = frozenset(
    {
        Column.PARTICIPANT_ID,
        Column.USERNAME,
        Column.APPLICATION_LABEL,
        Column.INTERACTION_TYPE,
        Column.APP_PACKAGE_NAME,
        Column.EVENT_TIMESTAMP,
    }
)


def _generate_plots(*args: Any, **kwargs: Any) -> Any:
    from chronicle_preprocessing_app.core.plotting.plotting_manager import generate_plots

    return generate_plots(*args, **kwargs)


@dataclass
class CellFormatRule:
    """Rule for formatting Excel cells based on a condition."""

    condition: Callable[[int, str, Any], bool]
    fill_color: str | None = None
    horizontal_alignment: str | None = None
    vertical_alignment: str | None = None

    def apply(self, cell: Any) -> None:
        from openpyxl.styles import Alignment, PatternFill

        if self.fill_color:
            cell.fill = PatternFill(
                start_color=self.fill_color,
                end_color=self.fill_color,
                fill_type="solid",
            )
        if self.horizontal_alignment or self.vertical_alignment:
            cell.alignment = Alignment(
                horizontal=self.horizontal_alignment or "general",
                vertical=self.vertical_alignment or "bottom",
            )


def write_df_to_excel_and_format(
    df: pl.DataFrame,
    save_path: Path | str,
    sheet_name: str,
    *,
    irregular_value_strategy: Callable[[int, str, Any], bool] | None = None,
    additional_format_rules: list[CellFormatRule] | None = None,
    if_sheet_exists: str | None = None,
) -> None:
    """Write a Polars dataframe to Excel and apply lightweight formatting."""
    import openpyxl

    save_path = Path(save_path)
    workbook = openpyxl.load_workbook(save_path) if save_path.exists() else openpyxl.Workbook()
    if sheet_name in workbook.sheetnames:
        if if_sheet_exists == "replace":
            del workbook[sheet_name]
        else:
            del workbook[sheet_name]
    active = workbook.active
    assert active is not None, "Workbook has no active sheet"
    is_blank = active.max_row == 1 and active.max_column == 1 and active["A1"].value is None
    worksheet = active if is_blank else workbook.create_sheet(sheet_name)
    worksheet.title = sheet_name

    rows = [df.columns, *df.iter_rows()]
    for row_index, row_values in enumerate(rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if row_index == 1:
                continue
            column_name = str(df.columns[column_index - 1])
            rules = [
                CellFormatRule(
                    condition=lambda _r, _c, _v: True,
                    horizontal_alignment="center",
                    vertical_alignment="center",
                )
            ]
            if irregular_value_strategy:
                rules.append(
                    CellFormatRule(
                        condition=lambda r, c, v: irregular_value_strategy(r, c, v),
                        fill_color="FFFF00",
                    )
                )
            if additional_format_rules:
                rules.extend(additional_format_rules)
            for rule in rules:
                if rule.condition(row_index, column_name, value):
                    rule.apply(cell)

    workbook.save(save_path)
    workbook.close()


class ChronicleAndroidRawDataPreprocessor:
    """Polars-first file and folder preprocessor."""

    def __init__(
        self,
        options: PreprocessingOptions,
        progress_callback: Callable[[str, int, int], None] | None = None,
    ) -> None:
        self.options = options
        self.progress_callback = progress_callback
        self.stats = ProcessingStats()
        self.timestamp_processor = TimestampPreprocessor(options)
        self.timezone_processor = TimezonePreprocessor(options)
        self.column_processor = ColumnPreprocessor(options)
        self.app_filter_processor = AppFilterPreprocessor(options)
        self.app_usage_processor = AppUsagePreprocessor(options)
        self.screen_usage_processor = ScreenUsagePreprocessor(options)
        self.study_date_provider = StudyDateRangeProvider(study_date_map=options.study_date_map)
        self.fast_preprocessor = PolarsFastPathPreprocessor(
            options,
            app_codebook=read_app_codebook(options.app_codebook_path)
            if options.use_app_codebook
            else None,
            study_date_provider=self.study_date_provider,
        )
        self.current_participant_id = ""
        self.current_participant_raw_data_df = pl.DataFrame()
        self.current_participant_screen_usage_df = pl.DataFrame()

        if self.options.use_filter_file and not self.options.apps_to_filter_dict:
            self.options.apps_to_filter_dict = read_filter_file(self.options.filter_file)
        if (
            self.options.use_apps_forcing_screen_open_file
            and not self.options.apps_forcing_screen_open_dict
        ):
            self.options.apps_forcing_screen_open_dict = read_apps_forcing_screen_open_file(
                self.options.apps_forcing_screen_open_file
            )
        if self.options.use_background_apps_file and not self.options.background_apps_dict:
            self.options.background_apps_dict = read_background_apps_file(
                self.options.background_apps_file
            )

    def get_participant_id_from_data(self) -> str:
        if self.current_participant_raw_data_df.is_empty():
            raise ValueError("No data loaded")
        series = self.current_participant_raw_data_df.get_column(Column.PARTICIPANT_ID)
        return str(series[1 if len(series) > 1 else 0])

    @staticmethod
    def _validate_required_columns(df: pl.DataFrame) -> None:
        """Raise ValueError if any required raw input columns are absent."""
        found = set(df.columns)
        missing = sorted(REQUIRED_RAW_COLUMNS - found)
        if missing:
            raise ValueError(
                f"Missing required columns: [{', '.join(missing)}]. Found: [{', '.join(sorted(found))}]"
            )

    def remove_selected_interaction_types(self) -> None:
        self.current_participant_raw_data_df = (
            self.fast_preprocessor._remove_selected_interaction_types(
                self.current_participant_raw_data_df
            )
        )

    def finalize_and_save_preprocessed_data_df(
        self,
        raw_data_filename: str,
        pre_algo_event_timestamps: pl.Series | None = None,
    ) -> Path:
        preprocessed_data_save_folder = (
            Path(self.options.output_folder)
            / f"{self.options.study_name + ' ' + PREPROCESSED_FOLDER_SUFFIX}"
        )
        preprocessed_data_save_folder.mkdir(parents=True, exist_ok=True)

        if not self.current_participant_raw_data_df.is_empty():
            if (
                self.options.process_screen_usage_sessions
                and not self.options.process_app_usage_sessions
                and self.current_participant_raw_data_df.filter(
                    pl.col(Column.INTERACTION_TYPE) == str(InteractionType.SCREEN_USAGE)
                ).height
            ):
                output_df = self.current_participant_raw_data_df
                output_file_suffix = f"Screen Usage {PREPROCESSED_FILE_SUFFIX}"
            else:
                output_df = self.current_participant_raw_data_df
                output_file_suffix = PREPROCESSED_FILE_SUFFIX

            output_df = output_df.with_columns(
                pl.lit(self.options.study_name).alias(Column.STUDY_NAME)
            )
            stem = Path(raw_data_filename).stem.replace("Raw ", "")
            app_save_name = preprocessed_data_save_folder / f"{stem} {output_file_suffix}"
            formatted = self.fast_preprocessor._format_output_frame(
                output_df.select(
                    [
                        column
                        for column in self.fast_preprocessor._build_output_columns(output_df)
                        if column in output_df.columns
                    ]
                )
            )
            formatted.write_csv(app_save_name)

            if pre_algo_event_timestamps is not None and len(pre_algo_event_timestamps) > 0:
                sidecar_path = app_save_name.with_name(
                    app_save_name.stem + GAP_TIMESTAMPS_SIDECAR_SUFFIX
                )
                pl.DataFrame({Column.EVENT_TIMESTAMP: pre_algo_event_timestamps}).write_parquet(
                    sidecar_path
                )

        if (
            self.options.process_app_usage_sessions
            and self.options.process_screen_usage_sessions
            and not self.current_participant_screen_usage_df.is_empty()
        ):
            screen_save_name = (
                preprocessed_data_save_folder
                / f"{Path(raw_data_filename).stem.replace('Raw ', '')} Screen Usage {PREPROCESSED_FILE_SUFFIX}"
            )
            screen_df = self.current_participant_screen_usage_df.with_columns(
                pl.lit(self.options.study_name).alias(Column.STUDY_NAME)
            )
            formatted_screen = self.fast_preprocessor._format_output_frame(
                screen_df.select(
                    [
                        column
                        for column in self.fast_preprocessor._build_output_columns(screen_df)
                        if column in screen_df.columns
                    ]
                )
            )
            formatted_screen.write_csv(screen_save_name)

        return preprocessed_data_save_folder

    def preprocess_Chronicle_Android_raw_data_file(
        self,
        raw_data_file: Path | str,
    ) -> tuple[Path, bool, dict[str, Any] | None]:
        raw_path = Path(raw_data_file)
        t_start = time.monotonic()
        LOGGER.debug(
            "Starting %s",
            self.__class__.__name__,
            extra={"row_count": None, "file": raw_path.name},
        )
        try:
            if supports_polars_fast_path(
                self.options,
                survey_data_processor_available=False,
            ):
                result = self.fast_preprocessor.preprocess_raw_data_file(raw_path)
                self.current_participant_id = result.participant_id
                self.current_participant_raw_data_df = result.data
                output_folder = self.fast_preprocessor.save_preprocessed_output(
                    result.data,
                    raw_data_filename=raw_path.name,
                    output_folder=self.options.output_folder,
                    study_name=self.options.study_name,
                    pre_algo_event_timestamps=result.pre_algo_event_timestamps,
                )
                self.stats.mark_processed(raw_path)
                elapsed = time.monotonic() - t_start
                final_rows = len(self.current_participant_raw_data_df)
                LOGGER.info(
                    "Preprocessed %s: %d → %d rows in %.2fs",
                    raw_path.name,
                    len(result.data),
                    final_rows,
                    elapsed,
                )
                return output_folder, True, None

            if self.options.model_concurrent_usage or self.options.use_background_apps_file:
                flag_name = (
                    "model_concurrent_usage"
                    if self.options.model_concurrent_usage
                    else "use_background_apps_file"
                )
                message = (
                    f"{flag_name} is only supported on the Polars fast "
                    "path, but the current options route to the legacy path "
                    "(e.g. survey-data or study-date providers active). "
                    "(Background apps rely on the concurrent-usage split, which "
                    "is fast-path only.)"
                )
                if not self.options.allow_concurrent_usage_fallback:
                    raise ValueError(
                        f"{message} Refusing to silently ignore the flag and emit "
                        "non-concurrent output. Set allow_concurrent_usage_fallback="
                        "True to proceed without concurrent-usage modeling."
                    )
                LOGGER.warning(
                    "%s Proceeding without concurrent-usage modeling "
                    "(allow_concurrent_usage_fallback is set).",
                    message,
                )

            raw_df = pl.read_csv(raw_path, infer_schema_length=10000)
            original_rows = len(raw_df)
            self._validate_required_columns(raw_df)
            self.current_participant_raw_data_df = raw_df
            self.current_participant_id = self.get_participant_id_from_data()

            df = self.column_processor.correct_username_column(raw_df)
            # Canonicalize interaction-type spellings ("Move to Foreground",
            # "Unknown importance: N") exactly like the fast path does — without
            # this, devices logging the alternate vocabulary look like they have
            # no app usage at all, and the screen-session derivation can't see
            # Screen Interactive/Non-Interactive events.
            df = self.fast_preprocessor._rename_interaction_types(df)
            df = self.timestamp_processor.correct_timestamp_column(df, Column.EVENT_TIMESTAMP)
            df = self.timezone_processor.apply_timezone_handling(df, Column.EVENT_TIMESTAMP)
            if self.options.correct_duplicate_event_timestamps:
                df = self.timestamp_processor.unalign_duplicate_timestamps(
                    df, Column.EVENT_TIMESTAMP
                )
            df = self.study_date_provider.filter_data_to_study_dates(
                df,
                self.current_participant_id,
                Column.EVENT_TIMESTAMP,
            )
            df = self.timestamp_processor.mark_data_time_gaps(
                df,
                Column.EVENT_TIMESTAMP,
                Column.DATA_TIME_GAP_HOURS,
            )
            df = self.column_processor.create_additional_columns(
                df,
                self.fast_preprocessor._get_possible_device_model(df),
            )
            df = self.app_filter_processor.label_filtered_apps(df)

            if self.options.process_screen_usage_sessions:
                screen_df = self.screen_usage_processor.derive_screen_usage_sessions(df)
                self.current_participant_screen_usage_df = screen_df.filter(
                    pl.col(Column.INTERACTION_TYPE) == str(InteractionType.SCREEN_USAGE)
                )
                if not self.options.process_app_usage_sessions:
                    self.current_participant_raw_data_df = self.current_participant_screen_usage_df
                    output_folder = self.finalize_and_save_preprocessed_data_df(raw_path.name)
                    self.stats.mark_processed(raw_path)
                    elapsed = time.monotonic() - t_start
                    final_rows = len(self.current_participant_raw_data_df)
                    LOGGER.info(
                        "Preprocessed %s: %d → %d rows in %.2fs",
                        raw_path.name,
                        original_rows,
                        final_rows,
                        elapsed,
                    )
                    return output_folder, True, None

            pre_algo_ts: pl.Series | None = None
            if self.options.process_app_usage_sessions:
                pre_algo_ts = df.get_column(Column.EVENT_TIMESTAMP).drop_nulls()
                df = self.app_usage_processor.run_app_usage_algorithm(df)
                df = self.fast_preprocessor._enrich_with_app_codebook_data(df)
                df = self.app_usage_processor.add_app_usage_details(df)
                df = self.app_usage_processor.add_app_usage_flags(df)
                self.current_participant_raw_data_df = df
            else:
                self.current_participant_raw_data_df = df

            self.remove_selected_interaction_types()
            output_folder = self.finalize_and_save_preprocessed_data_df(raw_path.name, pre_algo_ts)
            self.stats.mark_processed(raw_path)
            elapsed = time.monotonic() - t_start
            final_rows = len(self.current_participant_raw_data_df)
            LOGGER.info(
                "Preprocessed %s: %d → %d rows in %.2fs",
                raw_path.name,
                original_rows,
                final_rows,
                elapsed,
            )
            LOGGER.debug(
                "Completed %s",
                self.__class__.__name__,
                extra={"rows_in": original_rows, "rows_out": final_rows},
            )
            return output_folder, True, None
        except NoAppUsageDataError:
            self.stats.mark_empty_file(raw_path.name)
            return Path(), False, None
        except Exception as exc:
            self.stats.mark_error(raw_path, str(exc))
            raise

    def preprocess_Chronicle_Android_raw_data_folder(
        self,
        plotting_started_callback: Callable[[], None] | None = None,
        plotting_completed_callback: Callable[[], None] | None = None,
    ) -> tuple[Path, ProcessingStats]:
        files = get_matching_files_from_folder(
            self.options.raw_data_folder,
            self.options.raw_data_file_regex_pattern,
            ignore_names=["Survey", "Archive", "Do Not Use", "Preprocessed", PLOTTED_FOLDER_SUFFIX],
        )
        self.stats.total_files = len(files)
        if not files:
            return Path(), self.stats

        preprocessed_data_save_folder = Path()
        if self.options.parallel_processing and len(files) > 1:
            results, parallel_stats = preprocess_files_parallel(
                files,
                self.options,
                max_workers=self.options.parallel_max_workers,
                progress_callback=self.progress_callback,
            )
            _merge_processing_stats(self.stats, parallel_stats)
            for output_folder, success, _ in results:
                if success and output_folder != Path():
                    preprocessed_data_save_folder = output_folder
        else:
            for index, raw_data_file in enumerate(files, start=1):
                if self.progress_callback:
                    self.progress_callback(
                        f"Processing file {index}/{len(files)}: {raw_data_file.name}",
                        index,
                        len(files),
                    )
                output_folder, _, _ = self.preprocess_Chronicle_Android_raw_data_file(raw_data_file)
                if output_folder != Path():
                    preprocessed_data_save_folder = output_folder

        if self.options.enable_plotting and preprocessed_data_save_folder != Path():
            if plotting_started_callback:
                plotting_started_callback()
            try:
                _generate_plots(
                    study_name=self.options.study_name,
                    preprocessed_folder=preprocessed_data_save_folder,
                    options=self.options,
                    codebook_path=self.options.app_codebook_path,
                    progress_callback=self.progress_callback,
                )
            finally:
                if plotting_completed_callback:
                    plotting_completed_callback()

        return preprocessed_data_save_folder, self.stats


MainPreprocessor = ChronicleAndroidRawDataPreprocessor


def _process_single_file_worker(
    args: tuple[int, str, dict[str, Any]],
) -> tuple[int, Path, bool, dict[str, Any] | None, str, ProcessingStats]:
    input_index, file_path, options_dict = args
    options = PreprocessingOptions(**options_dict)
    preprocessor = ChronicleAndroidRawDataPreprocessor(options)
    try:
        output_folder, success, compliance_dict = (
            preprocessor.preprocess_Chronicle_Android_raw_data_file(Path(file_path))
        )
        return (
            input_index,
            output_folder,
            success,
            compliance_dict,
            Path(file_path).name,
            preprocessor.stats,
        )
    except Exception as exc:
        preprocessor.stats.mark_error(Path(file_path), str(exc))
        return input_index, Path(), False, None, Path(file_path).name, preprocessor.stats


def _merge_processing_stats(destination: ProcessingStats, source: ProcessingStats) -> None:
    destination.processed_files += source.processed_files
    destination.failed_files += source.failed_files
    destination.empty_files += source.empty_files
    destination.warnings.update(source.warnings)
    destination.errors.update(source.errors)
    destination.file_errors.update(source.file_errors)
    destination.processed_file_paths.update(source.processed_file_paths)


def _resolve_parallel_max_workers(max_workers: int | None, file_count: int) -> int:
    if max_workers is None or max_workers <= 0:
        max_workers = max(1, multiprocessing.cpu_count() // 2)
    if file_count > 0:
        max_workers = min(max_workers, file_count)
    return max(1, max_workers)


def _parallel_option_value(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, list):
        return list(value)
    if isinstance(value, set):
        return set(value)
    if isinstance(value, tuple):
        return tuple(value)
    return value


def _build_parallel_options_dict(options: PreprocessingOptions) -> dict[str, Any]:
    worker_options = {
        field.name: _parallel_option_value(getattr(options, field.name))
        for field in fields(PreprocessingOptions)
        if field.name != "survey_data_df"
    }
    worker_options["enable_plotting"] = False
    worker_options["parallel_processing"] = False
    worker_options["parallel_max_workers"] = None
    return worker_options


def preprocess_files_parallel(
    files: list[Path],
    options: PreprocessingOptions,
    max_workers: int | None = None,
    progress_callback: Callable[[str, int, int], None] | None = None,
) -> tuple[list[tuple[Path, bool, dict[str, Any] | None]], ProcessingStats]:
    worker_count = _resolve_parallel_max_workers(max_workers, len(files))
    options_dict = _build_parallel_options_dict(options)
    stats = ProcessingStats(total_files=len(files))
    results: list[tuple[Path, bool, dict[str, Any] | None]] = [None] * len(files)  # type: ignore[list-item]

    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(
                _process_single_file_worker,
                (index, str(file_path), options_dict),
            ): index
            for index, file_path in enumerate(files)
        }
        completed = 0
        for future in as_completed(futures):
            (
                input_index,
                output_folder,
                success,
                compliance_dict,
                file_name,
                worker_stats,
            ) = future.result()
            results[input_index] = (output_folder, success, compliance_dict)
            _merge_processing_stats(stats, worker_stats)
            completed += 1
            if progress_callback:
                progress_callback(
                    f"Processed file {completed}/{len(files)}: {file_name}",
                    completed,
                    len(files),
                )

    return results, stats
